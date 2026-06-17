#!/usr/bin/env python3
"""
create_template.py
------------------
Generates granny_template.xlsx — a pre-configured spreadsheet template for
the Granny genealogy app.

Usage:
    pip install openpyxl
    python create_template.py

Then upload granny_template.xlsx to Google Drive:
    Drive → New → File upload → select granny_template.xlsx
    Right-click the uploaded file → Open with → Google Sheets
    File → Save as Google Sheets

The resulting Google Sheet is ready to use with Code.gs — no manual
tab creation or header typing needed.
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Schema: tab name → ordered list of column headers
# Must exactly match the SHEETS constant in Code.gs
# ---------------------------------------------------------------------------

SCHEMA = {
    "Users": [
        "user_id",
        "email",
        "display_name",
        "role",
        "created_at",
    ],
    "People": [
        "person_id",
        "first_name",
        "middle_name",
        "last_name",
        "sex",
        "birth_date",
        "birth_place",
        "death_date",
        "death_place",
        "is_living",
        "notes",
        "created_by",
        "created_at",
        "updated_at",
    ],
    "Families": [
        "family_id",
        "spouse1_id",
        "spouse2_id",
        "union_type",
        "union_date",
        "union_place",
        "union_end_date",
        "union_end_reason",
        "notes",
        "created_by",
        "created_at",
        "updated_at",
    ],
    "Family_Children": [
        "family_id",
        "child_id",
        "relationship_type",
        "notes",
    ],
    "Events": [
        "event_id",
        "event_type",
        "event_date",
        "event_place",
        "title",
        "description",
        "source_citation",
        "created_by",
        "created_at",
        "updated_at",
    ],
    "Event_Participants": [
        "event_id",
        "person_id",
        "role",
        "notes",
    ],
}

# ---------------------------------------------------------------------------
# Dropdown validation: sheet → { column_name → comma-separated options }
# Applied to all data rows (2:1000).
# ---------------------------------------------------------------------------

DROPDOWNS = {
    "Users": {
        "role": "admin,editor,viewer",
    },
    "People": {
        "sex": "M,F,U",
        "is_living": "TRUE,FALSE",
    },
    "Families": {
        "union_type": "married,partnered,unknown",
        "union_end_reason": "divorce,death,annulment",
    },
    "Family_Children": {
        "relationship_type": "biological,adopted,step,foster",
    },
    "Events": {
        "event_type": "birth,death,marriage,census,immigration,military,graduation,residence,other",
    },
    "Event_Participants": {
        "role": "subject,spouse,child,parent,witness,informant,participant",
    },
}

# ---------------------------------------------------------------------------
# Approximate column widths (characters) — overridden per column where useful
# ---------------------------------------------------------------------------

WIDTH_DEFAULTS = {
    # IDs
    "user_id": 10, "person_id": 10, "family_id": 10, "event_id": 10,
    "child_id": 10, "spouse1_id": 12, "spouse2_id": 12,
    "created_by": 10,
    # Names / text
    "first_name": 16, "middle_name": 16, "last_name": 16,
    "display_name": 20, "email": 28,
    "title": 36, "description": 36, "notes": 36, "source_citation": 36,
    # Dates / places
    "birth_date": 14, "death_date": 14, "union_date": 14,
    "union_end_date": 14, "event_date": 14,
    "birth_place": 22, "death_place": 22, "union_place": 22, "event_place": 22,
    # Timestamps
    "created_at": 22, "updated_at": 22,
    # Short enums
    "sex": 8, "role": 12, "union_type": 12, "union_end_reason": 14,
    "relationship_type": 16, "event_type": 14,
    "is_living": 10,
}
DEFAULT_WIDTH = 18

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

# Header row: indigo-600 background (#4f46e5), white bold text
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

# Tab colours (hex, no #) — one per sheet for quick visual scanning
TAB_COLORS = {
    "Users":             "818CF8",  # indigo-400
    "People":            "34D399",  # emerald-400
    "Families":          "F472B6",  # pink-400
    "Family_Children":   "FB923C",  # orange-400
    "Events":            "60A5FA",  # blue-400
    "Event_Participants":"A78BFA",  # violet-400
}

HEADER_ROW_HEIGHT = 22  # points
DATA_ROW_HEIGHT   = 18  # points (applied to rows 2–10 as a visual hint)

# Thin border for the header bottom edge
_thin = Side(style="thin", color="3730A3")
HEADER_BORDER = Border(bottom=_thin)


def style_header_cell(cell):
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border    = HEADER_BORDER


def add_sheet(wb: openpyxl.Workbook, tab_name: str, headers: list[str]) -> None:
    ws = wb.create_sheet(title=tab_name)

    # Tab colour
    ws.sheet_properties.tabColor = TAB_COLORS.get(tab_name, "94A3B8")

    # Write and style header row
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header_cell(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            WIDTH_DEFAULTS.get(col_name, DEFAULT_WIDTH)
        )

    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Hint at row height for first few data rows
    for row in range(2, 8):
        ws.row_dimensions[row].height = DATA_ROW_HEIGHT

    # Freeze the header row so it stays visible while scrolling
    ws.freeze_panes = "A2"

    # Dropdown validation for constrained columns
    sheet_dropdowns = DROPDOWNS.get(tab_name, {})
    for col_name, options in sheet_dropdowns.items():
        if col_name not in headers:
            continue
        col_idx   = headers.index(col_name) + 1          # 1-based
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}1000"   # all data rows

        dv = DataValidation(
            type="list",
            formula1=f'"{options}"',
            allow_blank=True,
            showDropDown=False,   # False = show the dropdown arrow in the cell
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Must be one of: {options}",
        )
        dv.sqref = cell_range
        ws.add_data_validation(dv)

    # Auto-filter on header row (lets users sort/filter in Google Sheets)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"


def build_workbook() -> openpyxl.Workbook:
    """Return a fully configured Workbook — all 6 tabs, headers, styles, validation."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default blank sheet
    for tab_name, headers in SCHEMA.items():
        add_sheet(wb, tab_name, headers)
    return wb


def main():
    wb = build_workbook()
    output_path = "granny_template.xlsx"
    wb.save(output_path)
    print(f"Saved {output_path}")
    print()
    print("Next steps:")
    print("  1. Upload granny_template.xlsx to Google Drive")
    print("  2. Right-click → Open with → Google Sheets")
    print("  3. File → Save as Google Sheets")
    print("  4. Open Extensions → Apps Script and add Code.gs + Index.html")
    print()
    print("  Or run deploy.py to automate steps 1-4.")


if __name__ == "__main__":
    main()
